import os

from bip_utils import *
from hdwallet.utils import generate_mnemonic
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook import Workbook

INVERTED = "\033[7m"
RED = "\033[31m"
GREEN = "\033[32m"
RESET = "\033[0m"


def generate_wallets(amount: int):
    result = []

    for _ in range(amount):
        mnemonic = generate_mnemonic()
        seed = Bip39SeedGenerator(mnemonic).Generate("")

        bip44_mst_ctx = Bip44.FromSeed(seed, Bip44Coins.SOLANA)
        bip44_acc_ctx = bip44_mst_ctx.Purpose().Coin().Account(0)
        bip44_chg_ctx = bip44_acc_ctx.Change(Bip44Changes.CHAIN_EXT)

        data = {
            "address": bip44_chg_ctx.PublicKey().ToAddress(),
            "mnemonic": mnemonic
        }
        result.append(data)

    return result


def main():
    amount = None
    while amount is None:
        try:
            a = int(input("Укажите желаемое кол-во кошельков: "))
        except ValueError:
            print(f"{INVERTED}Неверное значение!{RESET} {RED}Повторите попытку.{RESET}\n")
            continue

        amount = a

    try:
        os.mkdir(os.getcwd() + os.sep + "output")
    except FileExistsError:
        pass

    filename = ""
    while filename == "":
        name = input("Укажите имя .xlsx-файла без расширения, куда будут записаны результаты: ")

        path = os.getcwd() + os.sep + "output" + os.sep + f"{name}.xlsx"
        if os.path.exists(path):
            print("Файл уже существует! Укажите уникальное имя.")
            continue
        else:
            filename = path

    wallets = generate_wallets(amount)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Solana"

    A1 = sheet.cell(row=1, column=1)
    A1.value = "Адрес"

    B1 = sheet.cell(row=1, column=2)
    B1.value = "Мнемоника"

    B1.font = A1.font = Font(
        name="Arial",
        bold=True,
        size=16,
        color="FFFFFF"
    )

    B1.fill = A1.fill = PatternFill(
        fill_type="solid",
        start_color="000000",
        end_color="000000"
    )

    B1.alignment = A1.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    row = 2
    address_width_set = False
    mnemonic_width_set = False
    for wallet in wallets:
        address_cell = sheet.cell(row=row, column=1)
        address_cell.value = wallet["address"]
        address_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        address_cell.font = Font(
            name="Consolas",
            bold=True,
            size=12
        )

        if not address_width_set:
            letter = address_cell.column_letter
            sheet.column_dimensions[letter].width = len(address_cell.value) * 1.5
            address_width_set = True

        mnemonic_cell = sheet.cell(row=row, column=2)
        mnemonic_cell.value = wallet["mnemonic"]
        mnemonic_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        mnemonic_cell.font = Font(
            name="Consolas",
            bold=False,
            size=12
        )

        if not mnemonic_width_set:
            letter = mnemonic_cell.column_letter
            sheet.column_dimensions[letter].width = len(mnemonic_cell.value) * 2
            mnemonic_width_set = True

        row += 1

    workbook.save(filename)
    print(INVERTED + GREEN + "Кошельки успешно сгенерированы!" + RESET)


if __name__ == "__main__":
    main()
